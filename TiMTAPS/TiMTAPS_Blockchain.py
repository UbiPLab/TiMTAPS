import ATS
import TiMTAPS
import json
from web3 import Web3, HTTPProvider
import random
import datetime
import xlwt
import openpyxl
from openpyxl import load_workbook

#0: signer:    85940a97b3ed61c451adac56a2e13a3ed6f42796
#1: combiner:  e4af8e73ac895a46c249e89504ad161638f65275
#2: tracer:    b462a24a24535b988488186741a4cd1f186faa5c
#3: admitter   2fed93e45c4d4fe1de0af385ebc451ce9a7c7b8b

#字符分割
def split_string(s):
    mid = len(s) // 2
    return s[:mid], s[mid:]

def Setup(lamda, k, n, t, T):
    pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(lamda, k, n, t, T)
    return pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal
def Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk_ATS):
    with open('TiMTAPS_sol_TiMTAPS_constract.abi', 'r') as f:
        abi = json.load(f)

    with open('TiMTAPS_sol_TiMTAPS_constract.bin', 'r') as f:
        bytecode = f.read()

    # 连接到本地 Geth 节点
    w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))

    # 确保连接成功
    if not w3.is_connected():
        print("Failed to connect to Ethereum node")

    # 获取默认账户
    default_account = w3.eth.accounts[0]
    # 部署合约
    SignatureVerifier = w3.eth.contract(abi=abi, bytecode=bytecode)
    tx_hash = SignatureVerifier.constructor().transact({'from': default_account})
    tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
    contract_address = tx_receipt.contractAddress

    #print(f"Contract deployed at address: {contract_address}")
    # 创建合约实例
    contract = w3.eth.contract(address=contract_address, abi=abi)
    sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk_ATS)
    # 获取合约的字节码和 ABI
    sigma_cor_string = json.dumps(sigma_cor)
    # print(sigma_cor_string.__len__())
    sigma_HTLP_cor_string = json.dumps(sigma_HTLP_cor)
    com_Rz_string = json.dumps(com_Rz)

    w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))
    # contract = w3.eth.contract(address=contract_Address, abi=abi)
    account_signer = w3.eth.accounts[0]  # 使用signer账户
    with open('keystore\\UTC--2025-01-07T07-58-12.532906000Z--85940a97b3ed61c451adac56a2e13a3ed6f42796','r') as f:
        encrypted_key = f.read()
    signer_private_key = w3.eth.account.decrypt(encrypted_key, '123456')
    # 设置交易参数
    transaction1 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    #估计使用的gas数
    #estimate_gas = contract.functions.setsigma_cor(sigma_cor_string).estimate_gas({'from': account_signer})
    #print(estimate_gas)
    #将sigma_cor上链
    try:
        param = sigma_cor_string  # 替换为实际参数
        function_call = contract.functions.setsigma_cor(param)
        #print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call.build_transaction(transaction1)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        #print(tx_receipt)
        #检测能不能正确读取
        #current_value = contract.functions.getsigma_cor().call()
        #print('Current value:', current_value)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    # 估计使用的gas数
    # estimate_gas = contract.functions.setsigma_HTLP_cor(sigma_HTLP_cor_string).estimate_gas({'from': account, 'gas': 40000000})
    # print(estimate_gas)
    # 将sigma_HTLP_cor上链
    sigma_HTLP_cor_string1, sigma_HTLP_cor_string2 = split_string(sigma_HTLP_cor_string)
    transaction2 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    # latest_block = w3.eth.get_block('latest')
    # print(f"Current block gas limit: {latest_block['gasLimit']}")
    try:
        param1 = sigma_HTLP_cor_string1# 替换为实际参数
        function_call1 = contract.functions.setsigma_HTLP_cor1(param1)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx1 = function_call1.build_transaction(transaction2)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx1, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        #print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    transaction3 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    try:
        param2 = sigma_HTLP_cor_string2  # 替换为实际参数
        function_call2 = contract.functions.setsigma_HTLP_cor2(param2)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx2 = function_call2.build_transaction(transaction3)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx2, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    transaction4 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    try:
        param = com_Rz_string # 替换为实际参数
        function_call2 = contract.functions.setcom_Rz(param)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call2.build_transaction(transaction4)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")
    return contract

def Combine(Sid, sk_com, m, PK, S, pp, idc, contract):

    #starttime = datetime.datetime.now()
    w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))
    # contract = w3.eth.contract(address=contract_Address, abi=abi)
    account = w3.eth.accounts[0]  # 使用第一个账户
    account_combiner = w3.eth.accounts[1]  # 使用combiner账户
    with open('keystore\\UTC--2025-01-07T08-33-03.406560300Z--e4af8e73ac895a46c249e89504ad161638f65275', 'r') as f:
        encrypted_key = f.read()
    combiner_private_key = w3.eth.account.decrypt(encrypted_key, '123456')
    try:
        sigma_cor_string = contract.functions.getsigma_cor().call()
    except Exception as e:
        print('Error calling contract:', e)
    #endtime = datetime.datetime.now()
    #timethis = endtime - starttime
    #print("Combine的数据下载所需要的时间是："+str(timethis))

    #starttime = datetime.datetime.now()
    sigma_cor_list = json.loads(sigma_cor_string)
    signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor_list, PK, S, pp, idc)
    signature_TiMTAPS_string = json.dumps(signature_TiMTAPS)
    signature_TiMTAPS_string1, signature_TiMTAPS_string2 = split_string(signature_TiMTAPS_string)
    #endtime = datetime.datetime.now()
    #timethis = endtime - starttime
    #print("Combine的签名聚合所需要的时间是：" + str(timethis))

    #starttime = datetime.datetime.now()
    transaction1 = {
        'from': account_combiner,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_combiner),
    }
    try:
        param = signature_TiMTAPS_string1  # 替换为实际参数
        function_call2 = contract.functions.setTiMTAPS1(param)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call2.build_transaction(transaction1)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, combiner_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    transaction2 = {
        'from': account_combiner,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_combiner),
    }
    try:
        param = signature_TiMTAPS_string2  # 替换为实际参数
        function_call2 = contract.functions.setTiMTAPS2(param)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call2.build_transaction(transaction2)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, combiner_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")
    #endtime = datetime.datetime.now()
    #timethis = endtime - starttime
    #print("Combine的数据上传所需要的时间是：" + str(timethis))

def Combine_Data(Sid, sk_com, m, PK, S, pp, idc, contract, file_path, row, column):
    wb = load_workbook(file_path)
    ws1 = wb['Sheet3']
    ws2 = wb['Sheet4']
    ws3 = wb['Sheet5']
    starttime = datetime.datetime.now()
    w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))
    # contract = w3.eth.contract(address=contract_Address, abi=abi)
    account = w3.eth.accounts[0]  # 使用第一个账户
    account_combiner = w3.eth.accounts[1]  # 使用combiner账户
    with open('keystore\\UTC--2025-01-07T08-33-03.406560300Z--e4af8e73ac895a46c249e89504ad161638f65275', 'r') as f:
        encrypted_key = f.read()
    combiner_private_key = w3.eth.account.decrypt(encrypted_key, '123456')
    try:
        sigma_cor_string = contract.functions.getsigma_cor().call()
    except Exception as e:
        print('Error calling contract:', e)
    endtime = datetime.datetime.now()
    timethis = endtime - starttime
    thetime = timethis.total_seconds()
    ws1.cell(row=row, column=column, value=thetime)
    print("Combine的数据下载所需要的时间是："+str(timethis))

    starttime = datetime.datetime.now()
    sigma_cor_list = json.loads(sigma_cor_string)
    signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor_list, PK, S, pp, idc)
    signature_TiMTAPS_string = json.dumps(signature_TiMTAPS)
    signature_TiMTAPS_string1, signature_TiMTAPS_string2 = split_string(signature_TiMTAPS_string)
    endtime = datetime.datetime.now()
    timethis = endtime - starttime
    thetime = timethis.total_seconds()
    ws2.cell(row=row, column=column, value=thetime)
    print("Combine的签名聚合所需要的时间是：" + str(timethis))

    starttime = datetime.datetime.now()
    transaction1 = {
        'from': account_combiner,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_combiner),
    }
    try:
        param = signature_TiMTAPS_string1  # 替换为实际参数
        function_call2 = contract.functions.setTiMTAPS1(param)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call2.build_transaction(transaction1)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, combiner_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    transaction2 = {
        'from': account_combiner,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_combiner),
    }
    try:
        param = signature_TiMTAPS_string2  # 替换为实际参数
        function_call2 = contract.functions.setTiMTAPS2(param)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call2.build_transaction(transaction2)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, combiner_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")
    endtime = datetime.datetime.now()
    timethis = endtime - starttime
    thetime = timethis.total_seconds()
    ws3.cell(row=row, column=column, value=thetime)
    wb.save(file_path)
    print("Combine的数据上传所需要的时间是：" + str(timethis))


def Verify(pp, PK, m, pk_ATS, contract):
    # 验证签名
    try:
        signature_TiMTAPS_string1 = contract.functions.getTiMTAPS1().call()
        signature_TiMTAPS_string2 = contract.functions.getTiMTAPS2().call()
        signature_TiMTAPS_string = signature_TiMTAPS_string1 + signature_TiMTAPS_string2
    except Exception as e:
        print('Error calling contract:', e)
    signature_TiMTAPS_list = json.loads(signature_TiMTAPS_string)

    try:
        com_Rz_string = contract.functions.getcom_Rz().call()
    except Exception as e:
        print('Error calling contract:', e)
    com_Rz_list = json.loads(com_Rz_string)
    return TiMTAPS.Verify(pp, PK, m, signature_TiMTAPS_list, com_Rz_list, pk_ATS)

def Trace(pp, PK, sk_tra, sk_adm, m, contract):
    try:
        signature_TiMTAPS_string1 = contract.functions.getTiMTAPS1().call()
        signature_TiMTAPS_string2 = contract.functions.getTiMTAPS2().call()
        signature_TiMTAPS_string = signature_TiMTAPS_string1 + signature_TiMTAPS_string2
    except Exception as e:
        print('Error calling contract:', e)
    signature_TiMTAPS_list = json.loads(signature_TiMTAPS_string)

    try:
        com_Rz_string = contract.functions.getcom_Rz().call()
    except Exception as e:
        print('Error calling contract:', e)
    com_Rz_list = json.loads(com_Rz_string)
    return TiMTAPS.Trace(pp, PK, sk_tra, sk_adm, m, signature_TiMTAPS_list, com_Rz_list)

def Open(S_id, sk_adm, pp, contract):
    try:
        sigma_HTLP_cor_string1 = contract.functions.getsigma_HTLP_cor1().call()
        sigma_HTLP_cor_string2 = contract.functions.getsigma_HTLP_cor2().call()
        sigma_HTLP_cor_string = sigma_HTLP_cor_string1 + sigma_HTLP_cor_string2
    except Exception as e:
        print('Error calling contract:', e)
    sigma_HTLP_cor_list = json.loads(sigma_HTLP_cor_string)
    return TiMTAPS.Open(S_id, sigma_HTLP_cor_list,sk_adm,pp)

def test_notDelay():
    k, n, t, T, Sid, idc = 5, 100, 45, 100, 1, 2
    m = "123746"
    S = random.sample(range(0, n), t)
    S.sort()
    # S = [0, 1, 2, 4, 7]
    print("签名者的序列是" + str(S))
    pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
    pk = sk_tra[0]
    pki = pk[1]
    contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
    Combine(Sid, sk_com, m, PK, S, pp, idc, contract)
    flag = Verify(pp, PK, m, pk, contract)
    S_t = Trace(pp, PK, sk_tra, sk_adm, m, contract)
    S_t.sort()
    print("Verify是否正确" + str(flag))
    print("追踪到的序列是" + str(S_t))
    return S == S_t

def test_Delay():
    # # 获取合约的字节码和 ABI
    # with open('TiMTAPS_sol_TiMTAPS_constract.abi', 'r') as f:
    #     abi = json.load(f)
    #
    # with open('TiMTAPS_sol_TiMTAPS_constract.bin', 'r') as f:
    #     bytecode = f.read()
    #
    # # 连接到本地 Geth 节点
    # w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))
    #
    # # 确保连接成功
    # if not w3.is_connected():
    #     print("Failed to connect to Ethereum node")
    #
    # # 获取默认账户
    # default_account = w3.eth.accounts[0]
    # # 部署合约
    # SignatureVerifier = w3.eth.contract(abi=abi, bytecode=bytecode)
    # tx_hash = SignatureVerifier.constructor().transact({'from': default_account})
    # tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
    # contract_address = tx_receipt.contractAddress
    # print(f"Contract deployed at address: {contract_address}")
    # # 创建合约实例
    # contract = w3.eth.contract(address=contract_address, abi=abi)
    k, n, t, T, Sid, idc = 5, 10, 5, 100, 1, 2
    m = "123746"
    S = random.sample(range(1, n), t)
    S.sort()
    # S = [2, 5, 7, 8]
    print("签名者的序列是" + str(S))
    pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
    pk = sk_tra[0]
    pki = pk[1]
    contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
    R, z = Open(Sid, sk_adm, pp, contract)
    S_t = ATS.Trace(pk, m, R, z, ATS.g, n, pk[0])
    S_t.sort()
    print("追踪到的序列是" + str(S_t))
    return S == S_t

def setup_data():

    k, n, t, T, Sid, idc = 5, 10, 5, 100, 1, 2
    # 创建一个工作簿
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet1')
    #n的取值20i
    sheet.write(0,0,'Setup') #表头
    for j in range(1,31):
        sheet.write(0, j, j) #表头
    sheet.write(0,j+1,'以秒数为单位')
    for i in range(1,11):
        print("n的值为："+str(20*i))
        n = 20*i
        sheet.write(i, 0, 20*i)
        for j in range(30):
            starttime = datetime.datetime.now()
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第"+str(j)+"次测量数据为："+str(thetime))
            thetime = thetime.total_seconds()
            sheet.write(i, j+1, thetime)

    # 保存工作簿
    workbook.save('data\\setup_data.xls')
def sign_data():
    k, n, t, T, Sid, idc = 5, 10, 5, 100, 1, 2
    m = ""
    S = random.sample(range(1, n), t)
    S.sort()
    # S = [2, 5, 7, 8]
    print("签名者的序列是" + str(S))
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet1')
    # n的取值20i
    sheet.write(0, 0, 'Sign')  # 表头
    for j in range(1, 101):
        sheet.write(0, j, j)  # 表头
    sheet.write(0, j + 1, '以秒数为单位')
    sheet.write(0, j + 2, 'm长度单位为')
    for i in range(1, 11):
        print("m的长度为：" + str(50 * i))
        for ml in range(50*i*1000*8):
            m = m + str(random.randint(0,9))
        sheet.write(i, 0, 50 * i)
        for j in range(100):
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            starttime = datetime.datetime.now()
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            sheet.write(i, j + 1, thetime)

    # 保存工作簿
    workbook.save('data\\sign_data.xls')
def sign_data2():
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    m = "123456"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet1')
    # n的取值20i
    sheet.write(0, 0, 'Sign')  # 表头
    for j in range(1, 101):
        sheet.write(0, j, j)  # 表头
    sheet.write(0, j + 1, '以秒数为单位')
    sheet.write(0, j + 2, 'm长度单位为')
    for i in range(1, 4):
        print("t的长度为：" + str(5 * i))
        t = 5 * i
        sheet.write(i, 0, str(5 * i))
        S = random.sample(range(1, n), t)
        S.sort()
        # S = [2, 5, 7, 8]
        #print("签名者的序列是" + str(S))
        for j in range(100):
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            starttime = datetime.datetime.now()
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            sheet.write(i, j + 1, thetime)

    # 保存工作簿
    workbook.save('data\\sign_data2.xls')
def combine_data():
    file_path = 'data\\combine_data.xlsx'
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet1']
    # n的取值20i
    ws.cell(row = 1, column = 1, value = 'Combine')  # 表头
    for j in range(1, 101):
        ws.cell(row=1, column=j+1, value=j)
    ws.cell(row=1, column=j+2, value='以秒数为单位')
    wb.save(file_path)
    #sheet.write(0, j + 2, 'm长度单位为')
    for i in range(1,4):
        t = 5 * i
        #sheet.write(i, 0, str(5 * i))
        for k in range(2,11):
            wb = load_workbook(file_path)
            ws = wb['Sheet1']
            round = 10*k
            head = "(" + str(5 * i) + "," + str(round) + ")"
            row = k + 10 * (i - 1)
            ws.cell(row=row+1, column=1, value=head)
            wb.save(file_path)
            print("第" + str(row) + "行第" + str(0) + "列已经写入")
            for j in range(1,101):
                starttime = datetime.datetime.now()
                timethis = starttime-starttime
                for rou in range(round):
                    print("round"+str(round)+"rou"+str(rou))
                    S = random.sample(range(1, n), t)
                    S.sort()
                        # S = [2, 5, 7, 8]
                        # print("签名者的序列是" + str(S))
                    pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
                    pk = sk_tra[0]
                    pki = pk[1]
                    contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
                    starttime = datetime.datetime.now()
                    Combine(Sid, sk_com, m, PK, S, pp, idc, contract)
                    endtime = datetime.datetime.now()
                    thetime = endtime - starttime
                    timethis = timethis +thetime
                print("第" + str(j) + "次测量数据为：" + str(timethis))
                timethis = timethis.total_seconds()
                wb = load_workbook(file_path)
                ws = wb['Sheet1']
                ws.cell(row=k+10*(i-1)+1, column=j, value=timethis)
                wb.save(file_path)
                print("第" + str(k+10*(i-1)) + "行第" + str(j + 1) + "列已经写入")

    # 保存工作簿
def combine_data_alone(i,t):
    file_path = 'test.xlsx'
    k, n, T, Sid, idc = 5, 100, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet3']
    # n的取值20i
    ws.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 101):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    # sheet.write(0, j + 2, 'm长度单位为')
    # sheet.write(i, 0, str(5 * i))
    wb = load_workbook(file_path)
    ws = wb['Sheet3']
    head = "(" + str(t) + "," + str(n) + ")"
    ws.cell(row=i+1, column=1, value=head)
    wb.save(file_path)
    print("第" + str(i+1) + "行第" + str(1) + "列已经写入")
    for j in range(1, 11):
        print("t="+ str(t)+"round" + str(j))
        S = random.sample(range(1, n), t)
        S.sort()
        # S = [2, 5, 7, 8]
        print("签名者的序列是" + str(S))
        pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
        pk = sk_tra[0]
        pki = pk[1]
        contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
        starttime = datetime.datetime.now()
        Combine(Sid, sk_com, m, PK, S, pp, idc, contract)
        endtime = datetime.datetime.now()
        thetime = endtime - starttime
        print("第" + str(j) + "次测量数据为：" + str(thetime))
        thetime = thetime.total_seconds()
        wb = load_workbook(file_path)
        ws = wb['Sheet3']
        ws.cell(row=i + 1, column=j+1, value=thetime)
        wb.save(file_path)
        print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
def combine_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 100, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet2']
    # n的取值20i
    ws.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 101):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 20):
        t = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet2']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 11):
        for i in range(1, 20):
            t = 5 * i
            print("t="+ str(t)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            starttime = datetime.datetime.now()
            Combine(Sid, sk_com, m, PK, S, pp, idc, contract)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet2']
            ws.cell(row=i  + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
#测combine各个部分的时间看看哪里出了问题
def combine_data_test():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 100, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws1 = wb['Sheet3']
    ws2 = wb['Sheet4']
    ws3 = wb['Sheet5']
    # n的取值20i
    ws1.cell(row=1, column=1, value='Combine')  # 表头
    ws2.cell(row=1, column=1, value='Combine')  # 表头
    ws3.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 31):
        ws1.cell(row=1, column=j + 1, value=j)
        ws2.cell(row=1, column=j + 1, value=j)
        ws3.cell(row=1, column=j + 1, value=j)
    ws1.cell(row=1, column=j + 2, value='以秒数为单位')
    ws2.cell(row=1, column=j + 2, value='以秒数为单位')
    ws3.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    # sheet.write(0, j + 2, 'm长度单位为')
    for i in range(1, 11):
        t = 5*i
        # sheet.write(i, 0, str(5 * i))
        wb = load_workbook(file_path)
        ws1 = wb['Sheet3']
        ws2 = wb['Sheet4']
        ws3 = wb['Sheet5']
        head = "(" + str(t) + "," + str(n) + ")"
        ws1.cell(row=i + 1, column=1, value=head)
        ws2.cell(row=i + 1, column=1, value=head)
        ws3.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i+1) + "行第" + str(1) + "列已经写入")

        for j in range(1, 31):
            print("t="+ str(t)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            #starttime = datetime.datetime.now()
            Combine_Data(Sid, sk_com, m, PK, S, pp, idc, contract, file_path, i + 1, j+1)
            #endtime = datetime.datetime.now()
            #thetime = endtime - starttime
            #print("第" + str(j) + "次测量数据为：" + str(thetime))
            #thetime = thetime.total_seconds()
            # wb = load_workbook(file_path)
            # ws = wb['Sheet1']
            # ws.cell(row=i + 1, column=j+1, value=thetime)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
def test():
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet1']
    # n的取值20i
    ws.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 101):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
def verify_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 50, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet6']
    # n的取值20i
    ws.cell(row=1, column=1, value='Verify')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 9):
        n = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet6']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(2, 10):
            n = 5 * i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            Combine(Sid, sk_com, m, PK, S, pp, idc, contract)
            starttime = datetime.datetime.now()
            flag = Verify(pp, PK, m, pk, contract)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet6']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")

def trace_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet8']
    # n的取值20i
    ws.cell(row=1, column=1, value='Trace')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(6, 20):
        n = i
        wb = load_workbook(file_path)
        ws = wb['Sheet8']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1 - 5, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1 - 5) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(6, 20):
            n = i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            Combine(Sid, sk_com, m, PK, S, pp, idc, contract)
            flag = Verify(pp, PK, m, pk, contract)
            starttime = datetime.datetime.now()
            S_t = Trace(pp, PK, sk_tra, sk_adm, m, contract)

            endtime = datetime.datetime.now()
            print("追踪到的序列是:" + str(S_t))
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet8']
            ws.cell(row=i + 1 - 5, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1 - 5) + "行第" + str(j + 1) + "列已经写入")

def open_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 10, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet10']
    # n的取值20i
    ws.cell(row=1, column=1, value='open')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(2, 11):
        n = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet10']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1 - 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1-1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(2, 11):
            n = 5 * i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            contract = Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            starttime = datetime.datetime.now()
            R, z = Open(Sid, sk_adm, pp, contract)
            endtime = datetime.datetime.now()
            #S_t = ATS.Trace(pk, m, R, z, ATS.g, n, pk[0])
            #print("追踪到的序列是:" + str(S_t))
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet10']
            ws.cell(row=i + 1 - 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1 - 1) + "行第" + str(j + 1) + "列已经写入")

def Sign_Bdata(sk_ATS, m, S, pki, PK, Sid, pp, pk_ATS, file_path, sheet,i,j):
    string_sheet = 'Sheet'+str(sheet)
    string_sheet2 = 'Sheet' + str(sheet+1)
    wb = load_workbook(file_path)
    ws1 = wb[string_sheet]
    ws2 = wb[string_sheet2]
    starttime = datetime.datetime.now()
    with open('TiMTAPS_sol_TiMTAPS_constract.abi', 'r') as f:
        abi = json.load(f)
    with open('TiMTAPS_sol_TiMTAPS_constract.bin', 'r') as f:
        bytecode = f.read()
    # 连接到本地 Geth 节点
    w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))
    # 确保连接成功
    if not w3.is_connected():
        print("Failed to connect to Ethereum node")
    # 获取默认账户
    default_account = w3.eth.accounts[0]
    # 部署合约
    SignatureVerifier = w3.eth.contract(abi=abi, bytecode=bytecode)
    tx_hash = SignatureVerifier.constructor().transact({'from': default_account})
    tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
    contract_address = tx_receipt.contractAddress

    #print(f"Contract deployed at address: {contract_address}")
    # 创建合约实例
    contract = w3.eth.contract(address=contract_address, abi=abi)
    endtime = datetime.datetime.now()
    # S_t = ATS.Trace(pk, m, R, z, ATS.g, n, pk[0])
    # print("追踪到的序列是:" + str(S_t))
    thetime = endtime - starttime
    starttimen = datetime.datetime.now()
    sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk_ATS)
    endtimen = datetime.datetime.now()
    thetimen = endtimen - starttimen
    thetimen = thetimen.total_seconds()
    ws1.cell(row=i, column=j, value=thetimen)
    print("sheet1第" + str(i) + "行第" + str(j) + "列已经写入")
    starttime = datetime.datetime.now()
    # 获取合约的字节码和 ABI
    sigma_cor_string = json.dumps(sigma_cor)
    # print(sigma_cor_string.__len__())
    sigma_HTLP_cor_string = json.dumps(sigma_HTLP_cor)
    com_Rz_string = json.dumps(com_Rz)

    w3 = Web3(Web3.HTTPProvider('http://127.0.0.1:8545'))
    # contract = w3.eth.contract(address=contract_Address, abi=abi)
    account_signer = w3.eth.accounts[0]  # 使用signer账户
    with open('keystore\\UTC--2025-01-07T07-58-12.532906000Z--85940a97b3ed61c451adac56a2e13a3ed6f42796','r') as f:
        encrypted_key = f.read()
    signer_private_key = w3.eth.account.decrypt(encrypted_key, '123456')
    # 设置交易参数
    transaction1 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    #估计使用的gas数
    #estimate_gas = contract.functions.setsigma_cor(sigma_cor_string).estimate_gas({'from': account_signer})
    #print(estimate_gas)
    #将sigma_cor上链
    try:
        param = sigma_cor_string  # 替换为实际参数
        function_call = contract.functions.setsigma_cor(param)
        #print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call.build_transaction(transaction1)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        #print(tx_receipt)
        #检测能不能正确读取
        #current_value = contract.functions.getsigma_cor().call()
        #print('Current value:', current_value)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    # 估计使用的gas数
    # estimate_gas = contract.functions.setsigma_HTLP_cor(sigma_HTLP_cor_string).estimate_gas({'from': account, 'gas': 40000000})
    # print(estimate_gas)
    # 将sigma_HTLP_cor上链
    sigma_HTLP_cor_string1, sigma_HTLP_cor_string2 = split_string(sigma_HTLP_cor_string)
    transaction2 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    # latest_block = w3.eth.get_block('latest')
    # print(f"Current block gas limit: {latest_block['gasLimit']}")
    try:
        param1 = sigma_HTLP_cor_string1# 替换为实际参数
        function_call1 = contract.functions.setsigma_HTLP_cor1(param1)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx1 = function_call1.build_transaction(transaction2)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx1, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        #print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    transaction3 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    try:
        param2 = sigma_HTLP_cor_string2  # 替换为实际参数
        function_call2 = contract.functions.setsigma_HTLP_cor2(param2)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx2 = function_call2.build_transaction(transaction3)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx2, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

    transaction4 = {
        'from': account_signer,
        'gas': 1000000000,
        'gasPrice': w3.to_wei('0.1', 'gwei'),
        'nonce': w3.eth.get_transaction_count(account_signer),
    }
    try:
        param = com_Rz_string # 替换为实际参数
        function_call2 = contract.functions.setcom_Rz(param)
        # print(type(function_call))
        # 确保函数调用后使用 buildTransaction
        tx = function_call2.build_transaction(transaction4)
        # 签名交易
        signed_txn = w3.eth.account.sign_transaction(tx, signer_private_key)
        # 发送交易
        tx_hash = w3.eth.send_raw_transaction(signed_txn.raw_transaction)
        # 等待交易确认
        tx_receipt = w3.eth.wait_for_transaction_receipt(tx_hash)
        #print(f"交易成功，交易哈希: {tx_hash.hex()}")
        # print(tx_receipt)
    except AttributeError as e:
        print(f"Error calling contract: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")
    endtime = datetime.datetime.now()
    print("第" + str(j) + "次测量数据为：" + str(thetime))
    thetime = thetime + endtime - starttime
    thetime = thetime.total_seconds()
    ws2.cell(row=i, column=j, value=thetime)
    print("sheet2第" + str(i) + "行第" + str(j) + "列已经写入")
    wb.save(file_path)
    return contract

def blockchainSign_data_te():
    k, n, t, T, Sid, idc = 5, 100, 45, 100, 1, 2
    m = "123746"
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet1']
    ws2 = wb['Sheet2']
    ws.cell(row=1, column=1, value='BlockchainSignt')
    ws2.cell(row=1, column=1, value='noBlockchainSignt')
    for j in range(1,31):
        ws.cell(row=1, column=j+1, value=j)
        ws2.cell(row=1, column=j+1, value=j)
        print("第" + str(1) + "行第" + str(j+1) + "列已经写入")
    wb.save(file_path)
    # S = [0, 1, 2, 4, 7]
    for i in range(1,20):
        t = 5*i
        S = random.sample(range(0, n), t)
        S.sort()
        print("签名者的序列是" + str(S))
        pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = Setup(1, k, n, t, T)
        pk = sk_tra[0]
        pki = pk[1]
        for j in range(1,31):
            contract = Sign_Bdata(sk_ATS, m, S, pki, PK, Sid, pp, pki, file_path, 1, i+1, j+1)
if __name__ == '__main__':
    #blockchainSign_data_te()
    #test_notDelay()
    #test_Delay()
    #setup_data()
    #sign_data()
    #sign_data2()
    #trace_data_te()
    #open_data_te()
    #test()
    combine_data_test()
    # 连接到以太坊节点
    # w3 = Web3(Web3.HTTPProvider('http://localhost:8545'))
    #
    # # 获取最新区块
    # latest_block = w3.eth.get_block('latest')
    #
    # # 打印 gasLimit
    # print(f"当前区块的 gas limit: {latest_block['gasLimit']}")
    #combine_data_alone(9, 45)