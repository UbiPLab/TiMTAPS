import TiMTAPS
import datetime
from openpyxl import load_workbook
import random
import gc
import hashlib

def setup_data():
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet1']
    k, n, t, T, Sid, idc = 5, 1, 1, 100, 1, 2
    # n的取值20i
    ws.cell(row=1, column=1, value='Setup')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 21):
        n = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet1']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(1, 21):
            n = 5 * i
            print("n=" + str(n) + "round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            starttime = datetime.datetime.now()
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet1']
            ws.cell(row=i + 1, column=j + 1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1) + "行第" + str(j + 1) + "列已经写入")

def sign_datam():
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet11']
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    # n的取值20i
    ws.cell(row=1, column=1, value='Sign')  # 表头
    for j in range(1, 21):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    m = ""
    S = random.sample(range(1, n), t)
    S.sort()
    print("签名者的序列是" + str(S))
    for i in range(1, 11):
        m = ""
        print("m的长度为：" + str(50 * i))
        for ml in range(50*i*1000*8):
            m = m + "1"
        wb = load_workbook(file_path)
        ws = wb['Sheet11']
        ws.cell(row = i+1, column = 1, value = 50 * i)
        wb.save(file_path)
    for i in range(1, 11):
        for j in range(1, 11):
            m = ""
            for ml in range(50*i*1000*8):
                m = m + "1"
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            S = random.sample(range(1, n), t)
            starttime = datetime.datetime.now()
            sigma_cor, sigma_HTLP_cor, com_Rz  = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            thetime = thetime.total_seconds()
            print("m的长度"+str(m.__len__())+"第" + str(j) + "次测量数据为：" + str(thetime))
            wb = load_workbook(file_path)
            ws = wb['Sheet11']
            ws.cell(row=i + 1, column=j + 1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1) + "行第" + str(j + 1) + "列已经写入")
            gc.collect()
def sign_datahead():
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet5']
    ws.cell(row=1, column=1, value='Sign')  # 表头
    for j in range(1, 11):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 11):
        m = ""
        print("m的长度为：" + str(50 * i))
        for ml in range(50*i*1000*8):
            m = m + str(random.randint(0,9))
        wb = load_workbook(file_path)
        ws = wb['Sheet5']
        ws.cell(row = i+1, column = 1, value = 50 * i)
        wb.save(file_path)
def sign_datamone(i, m):
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet6']
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    for j in range(1, 11):
        S = random.sample(range(1, n), t)
        S.sort()
        pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
        pk = sk_tra[0]
        pki = pk[1]
        starttime = datetime.datetime.now()
        sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pki)
        endtime = datetime.datetime.now()
        thetime = endtime - starttime
        thetime = thetime.total_seconds()
        print("m的长度为：" + str(m.__len__()))
        print("第" + str(j) + "次测量数据为：" + str(thetime))
        wb = load_workbook(file_path)
        ws = wb['Sheet6']
        ws.cell(row=i + 1, column=j + 1, value=thetime)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(j + 1) + "列已经写入")

def combine_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 100, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet3']
    # n的取值20i
    ws.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 11):
        t = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet3']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for i in range(1, 11):
        for j in range(1, 31):
            t = 5 * i
            print("t="+ str(t)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            starttime = datetime.datetime.now()
            signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet3']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
def combine_data_te1():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 100, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet4']
    # n的取值20i
    ws.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(2, 20):
        n = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet4']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(2, 20):
            n = 5 * i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            starttime = datetime.datetime.now()
            signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet4']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
def verify_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 50, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet5']
    # n的取值20i
    ws.cell(row=1, column=1, value='Verify')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 9):
        n = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet5']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(2, 10):
            n = 5 * i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            starttime = datetime.datetime.now()
            signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
            starttime = datetime.datetime.now()
            flag = TiMTAPS.Verify(pp, PK, m, signature_TiMTAPS, com_Rz, pk)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet5']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
def verify_data_te1():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 100, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet6']
    # n的取值20i
    ws.cell(row=1, column=1, value='Verify')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 20):
        t = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet6']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(1, 20):
            t = 5 * i
            print("t="+ str(t)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
            starttime = datetime.datetime.now()
            flag = TiMTAPS.Verify(pp, PK, m, signature_TiMTAPS, com_Rz, pk)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet6']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i+1) + "行第" + str(j + 1) + "列已经写入")
def trace_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet7']
    # n的取值20i
    ws.cell(row=1, column=1, value='Trace')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(6, 20):
        n = i
        wb = load_workbook(file_path)
        ws = wb['Sheet7']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1 - 5, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1 - 5) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(6, 20):
            n = i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
            flag = TiMTAPS.Verify(pp, PK, m, signature_TiMTAPS, com_Rz, pk)
            starttime = datetime.datetime.now()
            S_t = TiMTAPS.Trace(pp, PK, sk_tra, sk_adm, m, signature_TiMTAPS, com_Rz)

            endtime = datetime.datetime.now()
            print("追踪到的序列是:" + str(S_t))
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet7']
            ws.cell(row=i + 1 - 5, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1 - 5) + "行第" + str(j + 1) + "列已经写入")
def trace_data_tiao():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    t_all = [3,4,5]
    #t = 2
    #t_list = [3,4,5]
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet10']
    # n的取值20i
    ws.cell(row=1, column=1, value='Trace')  # 表头
    for j in range(1, 11):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in t_all:
        for k in range(1, 12):
            t = i
            wb = load_workbook(file_path)
            ws = wb['Sheet10']
            head = "((" + str(t) + "," + str(n) + ")"+ "," + str(k*10) + ")"
            hang = i-3
            ws.cell(row=hang + 1 + k + 10*hang, column=1, value=head)
            wb.save(file_path)
            print("第" + str(hang + 1 + k + 10*hang) + "行第" + str(1) + "列已经写入"+ str(k)+"组")
    # sheet.write(0, j + 2, 'm长度单位为')
    for i in t_all:
        for k in range(10, 11):
            for j in range(1, 11):
                t = i
                hang = i-3
                print("t=" + str(t) + "round" + str(j) + "group" + str(k * 10))
                ss = datetime.datetime.now()
                timegroup = ss - ss
                for group in range(10 * k):
                    S = random.sample(range(1, n), t)
                    S.sort()
                    # S = [2, 5, 7, 8]
                    # S = [2, 5, 7, 8]
                    print("签名者的序列是" + str(S))
                    pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
                    pk = sk_tra[0]
                    pki = pk[1]
                    sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
                    signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
                    flag = TiMTAPS.Verify(pp, PK, m, signature_TiMTAPS, com_Rz, pk)
                    starttime = datetime.datetime.now()
                    S_t = TiMTAPS.Trace(pp, PK, sk_tra, sk_adm, m, signature_TiMTAPS, com_Rz)
                    print(S_t == S)
                    endtime = datetime.datetime.now()
                    timegroup = timegroup + endtime - starttime
                print("group" + str(k + 10) + "第" + str(j) + "次测量数据为：" + str(timegroup))
                timegroup = timegroup.total_seconds()
                wb = load_workbook(file_path)
                ws = wb['Sheet10']
                ws.cell(row=hang + 1 + k + 10 * hang, column=j + 1, value=timegroup)
                wb.save(file_path)
                print("第" + str(hang + 1 + k + 10 * hang) + "行第" + str(j + 1) + "列已经写入")
def trace_data_te1():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 20, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet8']
    # n的取值20i
    ws.cell(row=1, column=1, value='Trace')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 20):
        t = i
        wb = load_workbook(file_path)
        ws = wb['Sheet8']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(1, 20):
            t = i
            print("t="+ str(t)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
            flag = TiMTAPS.Verify(pp, PK, m, signature_TiMTAPS, com_Rz, pk)
            starttime = datetime.datetime.now()
            S_t = TiMTAPS.Trace(pp, PK, sk_tra, sk_adm, m, signature_TiMTAPS, com_Rz)

            endtime = datetime.datetime.now()
            print("追踪到的序列是:" + str(S_t))
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet8']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1) + "行第" + str(j + 1) + "列已经写入")
def open_data_te():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 10, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet9']
    # n的取值20i
    ws.cell(row=1, column=1, value='open')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(2, 11):
        n = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet9']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1 - 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1-1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(2, 11):
            n = 5 * i
            print("n="+ str(n)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            starttime = datetime.datetime.now()
            R, z = TiMTAPS.Open(Sid, sigma_HTLP_cor, sk_adm, pp)
            endtime = datetime.datetime.now()
            #S_t = ATS.Trace(pk, m, R, z, ATS.g, n, pk[0])
            #print("追踪到的序列是:" + str(S_t))
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet9']
            ws.cell(row=i + 1 - 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1 - 1) + "行第" + str(j + 1) + "列已经写入")
def open_data_te1():
    file_path = 'test.xlsx'
    k, n, t, T, Sid, idc = 5, 100, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet10']
    # n的取值20i
    ws.cell(row=1, column=1, value='open')  # 表头
    for j in range(1, 31):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in range(1, 20):
        t = 5 * i
        wb = load_workbook(file_path)
        ws = wb['Sheet10']
        head = "(" + str(t) + "," + str(n) + ")"
        ws.cell(row=i + 1, column=1, value=head)
        wb.save(file_path)
        print("第" + str(i + 1) + "行第" + str(1) + "列已经写入")
    # sheet.write(0, j + 2, 'm长度单位为')
    for j in range(1, 31):
        for i in range(1, 20):
            t = 5 * i
            print("t="+ str(t)+"round" + str(j))
            S = random.sample(range(1, n), t)
            S.sort()
            # S = [2, 5, 7, 8]
            print("签名者的序列是" + str(S))
            pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
            pk = sk_tra[0]
            pki = pk[1]
            sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
            starttime = datetime.datetime.now()
            R, z = TiMTAPS.Open(Sid, sigma_HTLP_cor, sk_adm, pp)
            endtime = datetime.datetime.now()
            #S_t = ATS.Trace(pk, m, R, z, ATS.g, n, pk[0])
            #print("追踪到的序列是:" + str(S_t))
            thetime = endtime - starttime
            print("第" + str(j) + "次测量数据为：" + str(thetime))
            thetime = thetime.total_seconds()
            wb = load_workbook(file_path)
            ws = wb['Sheet10']
            ws.cell(row=i + 1, column=j+1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1) + "行第" + str(j + 1) + "列已经写入")

def combine_data_te_tiao():
    file_path = 'test.xlsx'
    t_all = [5]
    k, n, t, T, Sid, idc = 5, 15, 5, 100, 1, 2
    #t = 2
    m = "123456"
    wb = load_workbook(file_path)
    ws = wb['Sheet4']
    # n的取值20i
    ws.cell(row=1, column=1, value='Combine')  # 表头
    for j in range(1, 11):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    for i in t_all:
        for k in range(1, 12):
            t = i
            wb = load_workbook(file_path)
            ws = wb['Sheet4']
            head = "((" + str(t) + "," + str(n) + ")"+ "," + str(k*10) + ")"
            hang = i-5
            ws.cell(row=hang + 1 + k + 10*hang, column=1, value=head)
            wb.save(file_path)
            print("第" + str(hang + 1 + k + 10*hang) + "行第" + str(1) + "列已经写入"+ str(k)+"组")
    # sheet.write(0, j + 2, 'm长度单位为')
    for i in t_all:
        for k in range(10,11):
            for j in range(1, 11):
                t = i
                hang = i-5
                print("t=" + str(t) + "round" + str(j)+"group"+str(k*10))
                ss = datetime.datetime.now()
                timegroup = ss - ss
                for group in range(10*k):
                    S = random.sample(range(1, n), t)
                    S.sort()
                     # S = [2, 5, 7, 8]
                    print("签名者的序列是" + str(S))
                    pp, sk_ATS, PK, sk_com, sk_tra, sk_adm, sk_ver_EIGamal = TiMTAPS.Setup(1, k, n, t, T)
                    pk = sk_tra[0]
                    pki = pk[1]
                    sigma_cor, sigma_HTLP_cor, com_Rz = TiMTAPS.Sign(sk_ATS, m, S, pki, PK, Sid, pp, pk)
                    starttime = datetime.datetime.now()
                    signature_TiMTAPS = TiMTAPS.Combine(Sid, sk_com, m, sigma_cor, PK, S, pp, idc)
                    endtime = datetime.datetime.now()
                    timegroup = timegroup + endtime - starttime
                print("group"+str(k+10)+"第" + str(j) + "次测量数据为：" + str(timegroup))
                timegroup = timegroup.total_seconds()
                wb = load_workbook(file_path)
                ws = wb['Sheet4']
                ws.cell(row=hang + 1 + k + 10*hang, column=j+1, value=timegroup)
                wb.save(file_path)
                print("第" + str(hang + 1 + k + 10*hang) + "行第" + str(j + 1) + "列已经写入")

def hash_datam():
    file_path = 'test.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Sheet10']
    q = 154134170538138396189964328785507020323
    # n的取值20i
    ws.cell(row=1, column=1, value='Sign')  # 表头
    for j in range(1, 21):
        ws.cell(row=1, column=j + 1, value=j)
    ws.cell(row=1, column=j + 2, value='以秒数为单位')
    wb.save(file_path)
    m = ""
    for i in range(1, 11):
        m = ""
        print("m的长度为：" + str(50 * i))
        for ml in range(50*i*1000*8):
            m = m + str(random.randint(0,9))
        wb = load_workbook(file_path)
        ws = wb['Sheet10']
        ws.cell(row = i+1, column = 1, value = 50 * i)
        wb.save(file_path)
    for i in range(1, 11):
        m = ""
        for ml in range(50*i*1000*8):
            m = m + str(random.randint(0,9))
        for j in range(1, 11):
            starttime = datetime.datetime.now()
            c = int(hashlib.sha256(m.encode()).hexdigest(), 16) % (q-1)
            endtime = datetime.datetime.now()
            thetime = endtime - starttime
            thetime = thetime.total_seconds()
            print("m的长度"+str(m.__len__())+"第" + str(j) + "次测量数据为：" + str(thetime))
            wb = load_workbook(file_path)
            ws = wb['Sheet10']
            ws.cell(row=i + 1, column=j + 1, value=thetime)
            wb.save(file_path)
            print("第" + str(i + 1) + "行第" + str(j + 1) + "列已经写入")
            gc.collect()



if __name__ == '__main__':
    #setup_data()
    sign_datam()
    #combine_data_te()
    #combine_data_te1()
    #verify_data_te()
    #verify_data_te1()
    #trace_data_te()
    #trace_data_te1()
    #open_data_te()
    #open_data_te1()
    #combine_data_te_tiao()
    #sign_datam()
    #trace_data_tiao()
    # sign_datahead()
    # i = 3
    # m = ''
    # for ml in range(50 * i * 1000 * 8):
    #     m = m + str(random.randint(0, 9))
    # sign_datamone(i, m)